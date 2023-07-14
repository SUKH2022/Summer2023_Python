from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Examples from https://openpyxl.readthedocs.io/en/stable/usage.html

# Create Workbook
wb = Workbook()

dest_filename = 'empty_book.xlsx'

ws1 = wb.active
ws1.title = "range names"

# Append rows to sheet with any collection
for row in range(1, 40):
     ws1.append(range(600))

# Add a sheet
ws2 = wb.create_sheet(title="Pi")

# Edit the new sheet
ws2['F5'] = 3.14

# Add a formula
ws2["A1"] = "=SUM(F5, 1)"

# Create another sheet & dynamically fill cells with the column letter for the current column
ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)

# Save file
wb.save(filename = dest_filename)